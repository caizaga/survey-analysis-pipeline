"""
crosstabs.py
Genera tablas de cruce (n, % por fila, % por columna) según la matriz de análisis.
Exporta todas las tablas a output/crosstabs.xlsx, una hoja por cruce.

Uso como módulo (desde el notebook):
    from src.crosstabs import generate_all
    generate_all(df_le, df_long_actividad, df_long_grupos)
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT_DIR = BASE_DIR / "output"

# ── Alias de columnas ──────────────────────────────────────────────────────
COL = {
    'genero'         : '1_género_colocar_desde_la_base_de_datos',
    'rango_etario'   : '2_rango_etario_seleccionar_desde_la_base_de_datos',
    'ciudad'         : '3_ciudad_escribir_ej_quito',
    'vivienda'       : '4_vivienda_y_habitabilidad_rural_referido_a_campo_y_periferias_de_ciudades_identificadas_como_tal_urbano_ciudades_y_centros_cantonales_o_provinciales',
    'ingreso'        : '6_ingreso_promedio_mensual_tomando_como_referencia_500_usd',
    'n_asegurados'   : '7_no_de_personas_aseguradas_en_el_núcleo_familiar_escribir_número_sin_espacios_ej_4',
    'necesidad'      : '10_necesidad_del_asegurado_a_en_acudir_a_chequeos_médicos',
    'experiencia'    : '11_experiencia_categoria',
    'percepcion_salud': '13_percepción_del_estado_de_salud_del_afiliado_a',
    'gasto_salud'    : '14_gasto_en_salud_tomando_en_cuenta_el_uso_del_plan_salud_red',
    'promedio_gasto' : '15_promedio_mensual_de_presupuesto_gastado_en_salud_en_el_núcleo_familiar',
    'otros_seguros'  : '16_disponibilidad_de_otros_seguros_para_acceso_a_salud',
    'percepcion_respaldo': '17_percepción_de_respaldo_en_acceso_a_la_salud_debido_a_la_afiliación_al_plan_salud_red_de_salud',
}

# Etiquetas cortas para títulos en las hojas
LABEL = {
    'genero'          : 'Género (P1)',
    'rango_etario'    : 'Rango etario (P2)',
    'ciudad'          : 'Ciudad (P3)',
    'vivienda'        : 'Vivienda y habitabilidad (P4)',
    'actividad'       : 'Actividad económica (P5)',
    'ingreso'         : 'Ingreso promedio mensual (P6)',
    'n_asegurados'    : 'Nº personas aseguradas (P7)',
    'grupos'          : 'Grupos atención prioritaria (P8)',
    'necesidad'       : 'Necesidad de chequeos médicos (P10)',
    'experiencia'     : 'Experiencia de uso del seguro (P11)',
    'percepcion_salud': 'Percepción de salud del afiliado (P13)',
    'gasto_salud'     : 'Gasto en salud (P14)',
    'promedio_gasto'  : 'Promedio mensual gasto en salud (P15)',
    'otros_seguros'   : 'Disponibilidad de otros seguros (P16)',
    'percepcion_respaldo': 'Percepción de respaldo (P17)',
}

# ── Variables multichoice (usan df_long) ─────────────────────────────────
MULTICHOICE = {'actividad', 'grupos'}

# ── Matriz de cruces: (fila, columna_banner, nombre_hoja) ─────────────────
CROSSES = [
    ('rango_etario',   'genero',           'Q2xQ1'),
    ('vivienda',       'rango_etario',     'Q4xQ2'),
    ('actividad',      'genero',           'Q5xQ1'),
    ('actividad',      'rango_etario',     'Q5xQ2'),
    ('actividad',      'promedio_gasto',   'Q5xQ15'),
    ('actividad',      'otros_seguros',    'Q5xQ16'),
    ('ingreso',        'genero',           'Q6xQ1'),
    ('n_asegurados',   'percepcion_salud', 'Q7xQ13'),
    ('grupos',         'genero',           'Q8xQ1'),
    ('necesidad',      'percepcion_salud', 'Q10xQ13'),
    ('gasto_salud',    'otros_seguros',    'Q14xQ16'),
    ('promedio_gasto', 'genero',           'Q15xQ1'),
    ('promedio_gasto', 'rango_etario',     'Q15xQ2'),
    ('promedio_gasto', 'ingreso',          'Q15xQ6'),
    ('otros_seguros',  'genero',           'Q16xQ1'),
    ('otros_seguros',  'rango_etario',     'Q16xQ2'),
    ('otros_seguros',  'percepcion_salud', 'Q16xQ13'),
]

# ── Colores ───────────────────────────────────────────────────────────────
COLOR_TITLE_N   = "1F4E79"  # azul oscuro  → n
COLOR_TITLE_ROW = "375623"  # verde oscuro → % fila
COLOR_TITLE_COL = "7B2C2C"  # rojo oscuro  → % columna
COLOR_HEADER    = "D9E1F2"  # azul claro   → cabeceras de tabla


# ── Helpers ───────────────────────────────────────────────────────────────

def _fmt_pct(val) -> str:
    """Convierte float a '45,3%' (coma decimal, sin espacio)."""
    try:
        return f"{val:.2f}".replace(".", ",") + "%"
    except (TypeError, ValueError):
        return str(val)


def _numeric_sort(df: pd.DataFrame) -> pd.DataFrame:
    """
    Si el índice parece numérico (ej. Q7: '1','2','10'), reordena numéricamente
    para evitar que '10' quede antes que '2' en orden alfabético.
    Mantiene la fila 'Total' al final si existe.
    """
    has_total = "Total" in df.index
    body = df.drop("Total") if has_total else df
    try:
        order = pd.to_numeric(body.index, errors="raise").argsort()
        body = body.iloc[order]
    except (ValueError, TypeError):
        pass  # índice no numérico → dejar orden original
    return pd.concat([body, df.loc[["Total"]]]) if has_total else body


def _crosstab_regular(df: pd.DataFrame, row_col: str, col_col: str):
    ct_n   = pd.crosstab(df[row_col], df[col_col], margins=True, margins_name="Total")
    ct_row = pd.crosstab(df[row_col], df[col_col], normalize="index") * 100
    ct_col = pd.crosstab(df[row_col], df[col_col], normalize="columns") * 100
    ct_n   = _numeric_sort(ct_n)
    ct_row = _numeric_sort(ct_row)
    ct_col = _numeric_sort(ct_col)
    return ct_n, ct_row, ct_col


def _freq_univariate(df: pd.DataFrame, col: str):
    """
    Tabla de frecuencia simple para una variable regular.
    Retorna (ct_n, ct_pct) — cada uno es un DataFrame de una sola columna 'Total'.
    """
    total = len(df)
    n = df[col].value_counts()
    ct_n   = pd.DataFrame({"Total": n})
    ct_pct = pd.DataFrame({"Total": n / total * 100})
    ct_n.loc["Total"]   = total
    ct_pct.loc["Total"] = 100.0
    ct_n   = _numeric_sort(ct_n)
    ct_pct = _numeric_sort(ct_pct)
    return ct_n, ct_pct


def _freq_univariate_multi(df_long: pd.DataFrame, total_respondents: int):
    """
    Tabla de frecuencia simple para una variable multichoice.
    % = menciones / total encuestados → puede sumar >100%.
    """
    n = df_long["categoria"].value_counts()
    ct_n   = pd.DataFrame({"Total": n})
    ct_pct = pd.DataFrame({"Total": n / total_respondents * 100})
    ct_n.loc["Total"]   = n.sum()
    ct_pct.loc["Total"] = n.sum() / total_respondents * 100
    return ct_n, ct_pct


def _crosstab_multi(df_long: pd.DataFrame, df_le: pd.DataFrame, col_col: str):
    """
    Cruza el df long (multichoice) con la variable banner del df principal.

    % por columna usa como base el nº de RESPONDENTES por grupo banner (no menciones),
    por lo que puede sumar >100% por columna — es la interpretación correcta para
    variables de opción múltiple: "del total de X, qué % seleccionó esta opción".
    """
    merged = df_long.merge(df_le[["id", col_col]], on="id", how="left")
    ct_n   = pd.crosstab(merged["categoria"], merged[col_col], margins=True, margins_name="Total")
    ct_row = pd.crosstab(merged["categoria"], merged[col_col], normalize="index") * 100

    # Base correcta para col%: respondentes únicos por grupo banner, no menciones
    mention_counts      = pd.crosstab(merged["categoria"], merged[col_col])
    respondents_per_grp = df_le[col_col].value_counts()
    ct_col = mention_counts.div(respondents_per_grp, axis=1) * 100

    return ct_n, ct_row, ct_col


def _write_block(ws, df: pd.DataFrame, start_row: int, title: str,
                 title_color: str, is_pct: bool = False) -> int:
    """
    Escribe un bloque (título + cabecera + datos) en la hoja ws.
    Retorna la siguiente fila disponible.
    """
    n_cols = len(df.columns) + 1  # +1 por el índice

    # ── Fila de título ──
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(bold=True, color="FFFFFF", size=11)
    title_cell.fill = PatternFill("solid", fgColor=title_color)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    # Merge title across all columns
    if n_cols > 1:
        ws.merge_cells(
            start_row=start_row, start_column=1,
            end_row=start_row,   end_column=n_cols
        )

    # ── Fila de cabeceras ──
    header_row = start_row + 1
    idx_name = df.index.name or "Categoría"
    h0 = ws.cell(row=header_row, column=1, value=idx_name)
    h0.font = Font(bold=True)
    h0.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    h0.alignment = Alignment(horizontal="left", wrap_text=True)

    for j, col_name in enumerate(df.columns, start=2):
        hc = ws.cell(row=header_row, column=j, value=str(col_name))
        hc.font = Font(bold=True)
        hc.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        hc.alignment = Alignment(horizontal="center", wrap_text=True)

    # ── Filas de datos ──
    for i, (idx_val, row) in enumerate(df.iterrows()):
        r = header_row + 1 + i
        idx_cell = ws.cell(row=r, column=1, value=str(idx_val))
        idx_cell.alignment = Alignment(horizontal="left", wrap_text=True)
        if str(idx_val) == "Total":
            idx_cell.font = Font(bold=True)

        for j, val in enumerate(row, start=2):
            if is_pct:
                cell_val = _fmt_pct(val)
            else:
                cell_val = int(val) if not pd.isna(val) else 0
                if str(idx_val) == "Total":
                    ws.cell(row=r, column=j).font = Font(bold=True)
            data_cell = ws.cell(row=r, column=j, value=cell_val)
            data_cell.alignment = Alignment(horizontal="center")
            if str(idx_val) == "Total":
                data_cell.font = Font(bold=True)

    next_row = header_row + 1 + len(df) + 2  # 2 filas en blanco de separación
    return next_row


def _set_col_widths(ws, df: pd.DataFrame, idx_width: int = 40, data_width: int = 14):
    ws.column_dimensions[ws.cell(row=1, column=1).column_letter].width = idx_width
    for j in range(2, len(df.columns) + 2):
        col_letter = ws.cell(row=1, column=j).column_letter
        ws.column_dimensions[col_letter].width = data_width


# ── Función principal ─────────────────────────────────────────────────────

def generate_all(
    df_le: pd.DataFrame,
    df_long_actividad: pd.DataFrame,
    df_long_grupos: pd.DataFrame,
    output_filename: str = "crosstabs.xlsx",
) -> str:
    """
    Genera todas las tablas de cruce y las guarda en output/crosstabs.xlsx.
    Retorna la ruta del archivo generado.
    """
    OUTPUT_DIR.mkdir(exist_ok=True)
    output_path = OUTPUT_DIR / output_filename

    wb = Workbook()
    wb.remove(wb.active)  # elimina la hoja vacía por defecto

    long_map = {
        "actividad": df_long_actividad,
        "grupos":    df_long_grupos,
    }

    # ── Frecuencias univariadas (todas las variables — hoja única) ───────
    print("── Frecuencias univariadas ──")
    total_respondents = len(df_le)

    ws_freq = wb.create_sheet(title="Frecuencias")
    ws_freq.column_dimensions["A"].width = 40
    ws_freq.column_dimensions["B"].width = 14
    start = 1

    for key, col_name in COL.items():
        label = LABEL.get(key, key)
        ct_n, ct_pct = _freq_univariate(df_le, col_name)
        ct_n.index.name = ct_pct.index.name = label

        title_cell = ws_freq.cell(row=start, column=1, value=f"Frecuencia: {label}")
        title_cell.font = Font(bold=True, size=12)
        start += 1
        start = _write_block(ws_freq, ct_n,   start, "Frecuencias absolutas (n)",   COLOR_TITLE_N,   is_pct=False)
        start = _write_block(ws_freq, ct_pct, start, "% del total de encuestados",   COLOR_TITLE_COL, is_pct=True)
        print(f"  {key:25}  {label}")

    for key, df_long in long_map.items():
        label = LABEL.get(key, key)
        ct_n, ct_pct = _freq_univariate_multi(df_long, total_respondents)
        ct_n.index.name = ct_pct.index.name = label

        title_cell = ws_freq.cell(row=start, column=1, value=f"Frecuencia: {label}")
        title_cell.font = Font(bold=True, size=12)
        start += 1
        start = _write_block(ws_freq, ct_n,   start, "Frecuencias absolutas (n)",                                    COLOR_TITLE_N,   is_pct=False)
        start = _write_block(ws_freq, ct_pct, start, "% del total de encuestados — puede sumar >100% (opción múltiple)", COLOR_TITLE_COL, is_pct=True)
        print(f"  {key:25}  {label}")

    # ── Tablas de cruce ───────────────────────────────────────────────────
    print("\n── Tablas de cruce ──")
    for row_key, col_key, sheet_name in CROSSES:
        col_col = COL[col_key]
        row_label = LABEL.get(row_key, row_key)
        col_label = LABEL.get(col_key, col_key)

        # ── Calcular crosstabs ──
        if row_key in MULTICHOICE:
            df_long = long_map[row_key]
            ct_n, ct_row, ct_col = _crosstab_multi(df_long, df_le, col_col)
        else:
            row_col = COL[row_key]
            ct_n, ct_row, ct_col = _crosstab_regular(df_le, row_col, col_col)

        # Nombrar índice con etiqueta corta
        ct_n.index.name   = row_label
        ct_row.index.name = row_label
        ct_col.index.name = row_label

        # ── Crear hoja ──
        ws = wb.create_sheet(title=sheet_name)
        _set_col_widths(ws, ct_n)

        # Título de hoja (fila 1)
        ws.row_dimensions[1].height = 20
        intro = ws.cell(row=1, column=1, value=f"{row_label}  ×  {col_label}")
        intro.font = Font(bold=True, size=12)
        start = 3  # deja fila 2 en blanco

        # ── Bloques ──
        start = _write_block(ws, ct_n,   start, "Frecuencias absolutas (n)",                 COLOR_TITLE_N,   is_pct=False)
        start = _write_block(ws, ct_row, start, "% por fila  (base = total fila)",            COLOR_TITLE_ROW, is_pct=True)
        col_note = " — puede sumar >100% por columna" if row_key in MULTICHOICE else ""
        start = _write_block(ws, ct_col, start, f"% por columna  (base = total respondentes){col_note}", COLOR_TITLE_COL, is_pct=True)

        print(f"  {sheet_name:12}  {row_label}  ×  {col_label}")

    wb.save(output_path)
    print(f"\n✓ Guardado en: {output_path}")
    return str(output_path)
